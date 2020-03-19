import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors

if len(sys.argv) < 2:
    print('Useage: python3 multi_plication_table.py [n]')
    sys.exit()

N = int(sys.argv[1])

wb = Workbook()
ws = wb.active

ft = Font(color=colors.RED, bold=True, italic=True)

for n in range(2, N + 2):
    ws.cell(row=n, column=1).value = int(n-1)
    ws.cell(row=n, column=1).font = ft
    ws.cell(row=1, column=n).value = int(n-1)
    ws.cell(row=1, column=n).font = ft

for row_of_cell in range(2, N + 2):
    for col_of_cell in range(2, N + 2):
        ws.cell(row=row_of_cell, column=col_of_cell).value = (row_of_cell-1) * (col_of_cell -1)

wb.save("multi_plication.xlsx")
